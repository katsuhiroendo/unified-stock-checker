import streamlit as st
import asyncio
import pandas as pd
import os
import datetime
from urllib.parse import urlparse
import sys
from openpyxl import load_workbook

# 共通モジュールのパスを追加
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(BASE_DIR, "modules"))

from shopee_logic import ShopeeStockChecker, auto_upload_shopee, auto_download_shopee, ShopeeLoginRequiredError
from ebay_logic import EbayStockChecker
import threading
import time
import json

# ページ設定
st.set_page_config(
    page_title="Unified Stock Checker Dashboard",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# スタイル設定
st.markdown("""
<style>
    .main { background-color: #f8f9fa; }
    .stButton>button { width: 100%; border-radius: 5px; height: 3em; font-weight: bold; }
    .shopee-btn { background-color: #ee4d2d !important; color: white !important; }
    .ebay-btn { background-color: #0064d2 !important; color: white !important; }
    .all-btn { background-color: #28a745 !important; color: white !important; }
</style>
""", unsafe_allow_html=True)

st.title("📦 Unified Stock Checker Dashboard")
st.markdown("---")

# セッション状態の初期化
if "results_shopee" not in st.session_state: st.session_state.results_shopee = []
if "results_ebay" not in st.session_state: st.session_state.results_ebay = []
if "output_path_shopee" not in st.session_state: st.session_state.output_path_shopee = None
if "is_running" not in st.session_state: st.session_state.is_running = False
if "platform" not in st.session_state: st.session_state.platform = "Shopee"
if "elapsed_time" not in st.session_state: st.session_state.elapsed_time = None
if "start_time" not in st.session_state: st.session_state.start_time = None

# --- 自動巡回用のグローバル状態 ---
if "AUTOMONITOR_STATE" not in globals():
    globals()["AUTOMONITOR_STATE"] = {
        "is_enabled": False,
        "interval_hours": 4.0,
        "next_run": None,
        "last_run": None,
        "logs": []
    }

# セッションステートとグローバル状態の同期
if "auto_status_init" not in st.session_state:
    st.session_state.auto_status = globals()["AUTOMONITOR_STATE"]
    st.session_state.auto_status_init = True
else:
    globals()["AUTOMONITOR_STATE"] = st.session_state.auto_status

# ログ出力用ヘルパー
def write_auto_log(message):
    now = datetime.datetime.now()
    log_file = os.path.join(BASE_DIR, "logs", f"auto_monitor_{now.strftime('%Y%m%d')}.log")
    timestamp = now.strftime("[%Y-%m-%d %H:%M:%S]")
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    # ログファイル書き出し
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"{timestamp} {message}\n")
    
    # 画面表示用にグローバル状態へ保存
    state = globals().get("AUTOMONITOR_STATE")
    if state:
        state["logs"].append(f"{now.strftime('%H:%M:%S')} {message}")
        if len(state["logs"]) > 100: state["logs"].pop(0)
    
    print(f"{timestamp} {message}")

def run_auto_monitor_cycle(interval_hours):
    """
    定期実行される本体の処理
    """
    write_auto_log("=== 自動監視タスク開始 ===")
    try:
        # 1. eBay 処理
        write_auto_log("[eBay] 同期中...")
        eb_checker = EbayStockChecker(dry_run=False, auto_relist=False)
        success, msg = eb_checker.sync_ebay_data()
        if success:
            items = []
            if os.path.exists(eb_checker.items_csv):
                with open(eb_checker.items_csv, "r", encoding="utf-8") as f:
                    import csv
                    reader = csv.DictReader(f); items = [(r["ebay_id"], r["supplier_url"]) for r in reader]
            
            if items:
                write_auto_log(f"[eBay] {len(items)} 件の在庫チェック中...")
                # Thread 内での asyncio 実行
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                results = loop.run_until_complete(eb_checker.run_stock_check_batch(items))
                
                removed_count = 0
                for r in results:
                    if r["status"] in ("在庫なし", "判定不能"):
                        msg_detail = f"{r['ebay_id']} ({r['status']} / {r.get('reason','')}) URL: {r.get('url','')}"
                        write_auto_log(f"[eBay] 取り下げ対象検知: {msg_detail}")
                        if eb_checker.end_ebay_item(r["ebay_id"]):
                            removed_count += 1
                write_auto_log(f"[eBay] チェック完了. {removed_count} 件を取り下げました。")
                
                # 1.2 eBay 復活チェック
                ended_items = []
                already_relisted = eb_checker.get_already_relisted_ids()
                if os.path.exists(eb_checker.ended_items_csv):
                    with open(eb_checker.ended_items_csv, "r", encoding="utf-8") as f:
                        import csv
                        reader = csv.DictReader(f)
                        for row in reader:
                            eid = row.get("item_id") or row.get("ebay_id")
                            url = row.get("supplier_url")
                            if eid and eid not in already_relisted:
                                ended_items.append((eid, url))
                
                if ended_items:
                    write_auto_log(f"[eBay] {len(ended_items)} 件の終了済み商品（復活待ち）をチェック中...")
                    revival_results = loop.run_until_complete(eb_checker.run_stock_check_batch(ended_items, check_revival=True))
                    relisted_count = 0
                    for r in revival_results:
                        if r["status"] == "在庫あり":
                            write_auto_log(f"[eBay] 復活検知！再出品中: {r['ebay_id']}")
                            success, new_id = eb_checker.relist_ebay_item(r["ebay_id"])
                            if success:
                                eb_checker.record_relisted(r["ebay_id"], new_id, r["url"])
                                relisted_count += 1
                    write_auto_log(f"[eBay] 復活処理完了. {relisted_count} 件を再出品しました。")
                
                loop.close()
        else:
            write_auto_log(f"[eBay] 同期に失敗しました: {msg}")

        # 2. Shopee 処理
        write_auto_log("[Shopee] 同期中...")
        sh_checker = ShopeeStockChecker(headless=True)
        
        loop_sh = asyncio.new_event_loop()
        asyncio.set_event_loop(loop_sh)
        
        async def do_shopee_auto():
            from playwright.async_api import async_playwright
            async with async_playwright() as p:
                browser = await p.chromium.launch_persistent_context(
                    user_data_dir=sh_checker.user_data_dir, 
                    headless=True,
                    args=["--disable-blink-features=AutomationControlled"]
                )
                page = browser.pages[0] if browser.pages else await browser.new_page()
                
                try:
                    write_auto_log("[Shopee] 最新データをダウンロード中...")
                    download_path = await auto_download_shopee(page, sh_checker.download_dir)
                    
                    write_auto_log(f"[Shopee] {download_path} の在庫チェック中...")
                    output_path, summary, _, res_list = await sh_checker.run_full_cycle(None, skip_price_update=True, existing_file=download_path)
                    
                    oos_items = [r for r in res_list if r["result"]["status"] in ("SOLD_OUT", "UNKNOWN")]
                    revival_items = [r for r in res_list if r["result"]["status"] == "IN_STOCK" and r.get("current_stock", 0) == 0]
                    
                    if oos_items or revival_items:
                        if oos_items:
                            write_auto_log(f"[Shopee] {len(oos_items)} 件が在庫なし/判定不能。")
                        if revival_items:
                            write_auto_log(f"[Shopee] {len(revival_items)} 件に復活を検知。")
                        
                        from openpyxl import load_workbook
                        wb = load_workbook(download_path); ws = wb.active
                        update_list = []
                        
                        # 在庫切れ処理
                        for r in oos_items:
                            status = r["result"]["status"]
                            url = r["result"]["url"]
                            write_auto_log(f" - [{status}] ID:{r['display_id']} URL:{url}")
                            update_list.append({
                                "ID": r["display_id"], 
                                "在庫あり": False, 
                                "row_num": r.get("row_num"),
                                "sku_col": r.get("sku_col")
                            })
                        
                        # 復活処理
                        for r in revival_items:
                            url = r["result"]["url"]
                            write_auto_log(f" - [復活] ID:{r['display_id']} URL:{url}")
                            update_list.append({
                                "ID": r["display_id"], 
                                "在庫あり": True, 
                                "row_num": r.get("row_num"),
                                "sku_col": r.get("sku_col")
                            })
                        
                        write_auto_log("[Shopee] 在庫更新中...")
                        final_path, _ = sh_checker.save_manual_results(wb, ws, update_list, skip_price_update=True)
                        
                        write_auto_log("[Shopee] アップロード中...")
                        success = await auto_upload_shopee(page, final_path)
                        if success:
                            write_auto_log("[Shopee] アップロード成功。")
                        else:
                            write_auto_log("[Shopee] アップロードに失敗しました。")
                    else:
                        write_auto_log("[Shopee] 在庫切れはありません。")
                except ShopeeLoginRequiredError:
                    write_auto_log("[Shopee] ログインが必要です。ヘッドレスをオフにして一度ログインしてください。")
                except Exception as e:
                    write_auto_log(f"[Shopee] 実行エラー: {e}")
                finally:
                    await browser.close()

        loop_sh.run_until_complete(do_shopee_auto())
        loop_sh.close()

    except Exception as e:
        write_auto_log(f"自動監視中にエラーが発生しました: {e}")
    
    write_auto_log("=== 自動監視タスク終了 ===")

def auto_monitor_worker():
    """
    バックグラウンドで無限ループするワーカー
    """
    state = globals()["AUTOMONITOR_STATE"]
    while True:
        if state["is_enabled"]:
            now = datetime.datetime.now()
            if not state["last_run"] or \
               now >= state["next_run"]:
                
                run_auto_monitor_cycle(state["interval_hours"])
                
                state["last_run"] = now
                state["next_run"] = now + datetime.timedelta(hours=state["interval_hours"])
                state["logs"].append(f"{now.strftime('%H:%M:%S')} 実行完了")
                if len(state["logs"]) > 10: state["logs"].pop(0)
        
        time.sleep(60) # 1分ごとにチェック

# スレッド起動 (重複起動を防ぐために名前でチェック)
def start_auto_monitor_thread():
    exists = any(t.name == "AutoMonitorThread" for t in threading.enumerate())
    if not exists:
        monitor_thread = threading.Thread(target=auto_monitor_worker, name="AutoMonitorThread", daemon=True)
        monitor_thread.start()
        write_auto_log("--- 自動監視スレッドを起動しました ---")

if "AUTOMONITOR_STATE" in globals():
    start_auto_monitor_thread()

# サイドバー
with st.sidebar:
    st.header("⚙️ 設定")
    platform = st.selectbox("対象プラットフォームを選択", ["Shopee", "eBay", "All (Shopee & eBay)"], key="platform_select")
    st.session_state.platform = platform
    
    st.markdown("---")
    headless = st.checkbox("ヘッドレスモード (ブラウザを表示しない)", value=True)
    
    skip_price_update = True
    dry_run = False
    auto_relist = False
    if "Shopee" in platform or "All" in platform:
        pass
    
    if "eBay" in platform or "All" in platform:
        auto_relist = st.checkbox("在庫復活時に自動再出品 (eBay)", value=False)
        
    st.markdown("---")
    st.header("🕒 自動監視モード")
    auto_enabled = st.checkbox("自動監視を有効にする", 
                                value=st.session_state.auto_status["is_enabled"],
                                help="チェックを入れると設定された周期で自動チェックと取り下げを実行します。")
    st.session_state.auto_status["is_enabled"] = auto_enabled
    
    interval = st.number_input("巡回周期（時間）", 
                                min_value=0.5, 
                                max_value=24.0, 
                                value=st.session_state.auto_status["interval_hours"], 
                                step=0.5)
    st.session_state.auto_status["interval_hours"] = interval
    
    if auto_enabled:
        if st.session_state.auto_status["next_run"]:
            st.info(f"⏭️ 次回予定: {st.session_state.auto_status['next_run'].strftime('%H:%M:%S')}")
        else:
            st.info("🔜 まもなく初回実行を開始します")
        
        st.markdown("#### 🛡️ 監視ログ")
        if st.session_state.auto_status["logs"]:
            log_text = "\n".join(st.session_state.auto_status["logs"])
            st.code(log_text, language="text")
        else:
            st.caption("実行待ち...")

    st.markdown("---")
    if st.button("🔌 アプリを終了する"):
        st.warning("プログラムを停止しました。")
        os._exit(0)

# メイン画面 (同時実行時は2カラムレイアウト)
st.subheader(f"🚀 {platform} 操作")
start_button = st.button(f"在庫チェックを開始", disabled=st.session_state.is_running)

# 経過時間表示エリア
elapsed_placeholder = st.empty()
if st.session_state.elapsed_time:
    elapsed_placeholder.success(f"⏱️ 前回の処理時間: {st.session_state.elapsed_time}")

# プレースホルダー作成
p_col1, p_col2 = st.columns(2)
with p_col1:
    st.markdown("### Shopee Status")
    shopee_status = st.empty()
    shopee_progress = st.empty()
    sm_col1, sm_col2, sm_col3, sm_col4 = st.columns(4)
    sm1, sm2, sm3, sm4 = sm_col1.empty(), sm_col2.empty(), sm_col3.empty(), sm_col4.empty()

with p_col2:
    st.markdown("### eBay Status")
    ebay_status = st.empty()
    ebay_progress = st.empty()
    em_col1, em_col2, em_col3, em_col4 = st.columns(4)
    em1, em2, em3, em4 = em_col1.empty(), em_col2.empty(), em_col3.empty(), em_col4.empty()

# 実行ロジック
async def run_shopee(checker, callback):
    try:
        path, summary, page, res_list = await checker.run_full_cycle(callback, skip_price_update=skip_price_update)
        st.session_state.output_path_shopee = path
        return True
    except ShopeeLoginRequiredError:
        st.error("🔒 Shopeeへのログインが必要です。ヘッドレスモードをオフにしてログインしてください。")
        return False
    except Exception as e:
        st.error(f"Shopeeエラー: {e}")
        return False

def run_ebay_sync(checker, callback, status_cb=None):
    try:
        if status_cb: status_cb("📡 eBay データを同期中...")
        success, msg = checker.sync_ebay_data()
        
        if not success:
            if status_cb: status_cb(f"❌ 同期失敗: {msg}")
            st.error(f"eBay同期エラー: {msg}")
            return False
        
        import csv
        # 1. 出品中アイテムのチェック
        active_items = []
        if os.path.exists(checker.items_csv):
            with open(checker.items_csv, mode="r", encoding="utf-8") as f:
                reader = csv.DictReader(f); active_items = [(r["ebay_id"], r["supplier_url"]) for r in reader]
        
        if status_cb: status_cb(f"🔍 {len(active_items)} 件のアクティブ在庫チェックを開始...")
        # Phase 2: Active items check (Parallel)
        results_active = asyncio.run(checker.run_stock_check_batch(active_items, callback=lambda r, c, t: callback(r, c, t, is_active=True)))
        
        # 2. 終了済みアイテムのチェック (在庫0のもの)
        ended_items = []
        already = checker.get_already_relisted_ids()
        if os.path.exists(checker.ended_items_csv):
            with open(checker.ended_items_csv, mode="r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    eid = row.get("item_id") or row.get("ebay_id")
                    url = row.get("supplier_url")
                    if eid and eid not in already: ended_items.append((eid, url))
        
        if ended_items:
            if status_cb: status_cb(f"🔄 {len(ended_items)} 件の終了済みアイテム（旧在庫0）をチェック中...")
            # 終了済みアイテムは is_active=False でコールバック
            asyncio.run(checker.run_stock_check_batch(ended_items, check_revival=True, callback=lambda r, c, t: callback(r, c, t, is_active=False)))

        if status_cb: status_cb("✅ eBay すべてのチェック完了")
        return True
    except Exception as e:
        if status_cb: status_cb(f"❌ エラー: {e}")
        st.error(f"eBayエラー: {e}")
        return False

if start_button:
    import time
    st.session_state.is_running = True
    st.session_state.results_shopee = []
    st.session_state.results_ebay = []
    st.session_state.output_path_shopee = None
    st.session_state.elapsed_time = None
    st.session_state.start_time = time.time()
    elapsed_placeholder.info("⏱️ 計測中...")

    async def start_all():
        tasks = []
        if "Shopee" in platform or "All" in platform:
            shopee_checker = ShopeeStockChecker(headless=headless)
            def shopee_cb(text, cur, tot, res=None):
                shopee_status.markdown(f"**状態:** {text}")
                if tot > 0: shopee_progress.progress(cur/tot)
                if res:
                    st.session_state.results_shopee = res
                    in_s = len([x for x in res if x["result"]["status"]=="IN_STOCK"])
                    sold = len([x for x in res if x["result"]["status"]=="SOLD_OUT"])
                    unkn = len([x for x in res if x["result"]["status"]=="UNKNOWN"])
                    sm1.metric("トータル", len(res)); sm2.metric("在庫あり", in_s); sm3.metric("在庫無し", sold, delta_color="inverse"); sm4.metric("判定不能", unkn)
            tasks.append(run_shopee(shopee_checker, shopee_cb))
        else:
            shopee_status.info("対象外")

        if "eBay" in platform or "All" in platform:
            from streamlit.runtime.scriptrunner import get_script_run_ctx, add_script_run_ctx
            ctx = get_script_run_ctx()
            ebay_checker = EbayStockChecker(dry_run=dry_run, auto_relist=auto_relist)
            
            def ebay_cb(res, cur, tot, is_active=True):
                ebay_status.markdown(f"**🔍 チェック中:** {cur}/{tot} | ID: `{res['ebay_id']}`")
                ebay_progress.progress(cur/tot)
                res_data = {
                    "display_id": res["ebay_id"], 
                    "is_active": is_active,
                    "result": {
                        "status": "IN_STOCK" if res["status"]=="在庫あり" else "SOLD_OUT" if res["status"]=="在庫なし" else "UNKNOWN", 
                        "url": res["url"], 
                        "price": 0,
                        "reason": res.get("reason", "")
                    }
                }
                st.session_state.results_ebay.append(res_data)
                r_list = st.session_state.results_ebay
                # トータル表示を全体の累積数に変更
                in_s_e = len([x for x in r_list if x["result"]["status"]=="IN_STOCK"])
                sold_s_e = len([x for x in r_list if x["result"]["status"]=="SOLD_OUT"])
                unkn_s_e = len([x for x in r_list if x["result"]["status"]=="UNKNOWN"])
                
                em1.metric("処理済み", len(r_list))
                em2.metric("在庫あり", in_s_e)
                em3.metric("在庫無し", sold_s_e, delta_color="inverse")
                em4.metric("判定不能", unkn_s_e)
            
            def ebay_status_cb(msg):
                ebay_status.markdown(f"**状態:** {msg}")
            
            def ebay_thread_wrapper():
                add_script_run_ctx(None, ctx)
                return run_ebay_sync(ebay_checker, ebay_cb, status_cb=ebay_status_cb)
            
            tasks.append(asyncio.to_thread(ebay_thread_wrapper))
        else:
            ebay_status.info("対象外")

        results = await asyncio.gather(*tasks)
        st.session_state.is_running = False
        # 経過時間を計算して保存
        import time
        elapsed_sec = time.time() - st.session_state.start_time
        minutes = int(elapsed_sec // 60)
        seconds = int(elapsed_sec % 60)
        if minutes > 0:
            st.session_state.elapsed_time = f"{minutes}分 {seconds}秒"
        else:
            st.session_state.elapsed_time = f"{seconds}秒"
        elapsed_placeholder.success(f"⏱️ 処理時間: {st.session_state.elapsed_time}")

    loop = asyncio.new_event_loop(); asyncio.set_event_loop(loop)
    loop.run_until_complete(start_all())

# 結果リスト表示 (タブ形式で整理)
if st.session_state.results_shopee or st.session_state.results_ebay:
    st.markdown("---")
    res_tab1, res_tab2 = st.tabs(["🛍️ Shopee Results", "🌎 eBay Results"])
    
    with res_tab1:
        if st.session_state.results_shopee:
            data_s = []
            for r in st.session_state.results_shopee:
                res = r["result"]
                data_s.append({
                    "row_num": r.get("row_num"), 
                    "sku_col": r.get("sku_col"), 
                    "ID": r["display_id"], 
                    "現在庫": r.get("current_stock", 0),
                    "在庫あり": (res["status"] == "IN_STOCK"), 
                    "販売サイト": res["url"], 
                    "URL": res["url"], 
                    "_status": res["status"]
                })
            df_s = pd.DataFrame(data_s)
            f_status_s = st.multiselect("Shopee表示フィルター", ["IN_STOCK", "SOLD_OUT", "UNKNOWN"], default=["IN_STOCK", "SOLD_OUT", "UNKNOWN"], key="f_s")
            disp_s = df_s[df_s["_status"].isin(f_status_s)] if f_status_s else df_s
            
            if not disp_s.empty:
                edited_s = st.data_editor(
                    disp_s, 
                    column_config={
                        "row_num": None, 
                        "sku_col": None,
                        "_status": None, 
                        "ID": st.column_config.TextColumn("ID", disabled=True), 
                        "現在庫": st.column_config.NumberColumn("現在庫", disabled=True),
                        "販売サイト": st.column_config.TextColumn("販売サイト", disabled=False),
                        "URL": st.column_config.LinkColumn("商品ページ")
                    }, 
                    width='stretch', 
                    hide_index=True, 
                    key="ed_s"
                )
            else:
                st.info("💡 フィルター条件に一致する結果はありません。")
            
            if st.session_state.output_path_shopee:
                if st.button("🚀 Shopeeへアップロード反映", type="primary"):
                    with st.spinner("反映中..."):
                        try:
                            from shopee_logic import auto_upload_shopee
                            checker = ShopeeStockChecker(headless=headless)
                            wb = load_workbook(st.session_state.output_path_shopee); ws = wb.active
                            final_path, _ = checker.save_manual_results(wb, ws, edited_s.to_dict('records'), skip_price_update=skip_price_update)
                            async def up():
                                from playwright.async_api import async_playwright
                                async with async_playwright() as p:
                                    browser = await p.chromium.launch_persistent_context(user_data_dir=checker.user_data_dir, headless=checker.headless, args=["--disable-blink-features=AutomationControlled"])
                                    up_page = browser.pages[0] if browser.pages else await browser.new_page()
                                    success = await auto_upload_shopee(up_page, final_path)
                                    await browser.close(); return success
                            
                            up_loop = asyncio.new_event_loop(); asyncio.set_event_loop(up_loop)
                            if up_loop.run_until_complete(up()): st.success("🎉 Shopeeアップロード完了！")
                        except Exception as e:
                            st.error(f"エラー: {e}")
        else: st.info("Shopeeの実行結果はありません。")

    with res_tab2:
        st.markdown("#### ⚙️ リスティング管理")
        st.caption("下表の「在庫あり」チェックボックスを調整して、反映ボタンを押してください。")
        st.caption("- チェックあり: 出品維持 / 再出品実行")
        st.caption("- チェックなし: 取り下げ実行 / 終了維持")
        
        # --- 結果テーブル ---
        if st.session_state.results_ebay:
            data_e = []
            for r in st.session_state.results_ebay:
                curr_stock = 1 if r.get("is_active", True) else 0 
                data_e.append({
                    "ID": r["display_id"], 
                    "現在庫": curr_stock,
                    "在庫あり": (r["result"]["status"]=="IN_STOCK"), 
                    "販売サイト": r["result"]["url"],
                    "URL": r["result"]["url"], 
                    "_status": r["result"]["status"]
                })
            df_e = pd.DataFrame(data_e)
            f_status_e = st.multiselect("eBay表示フィルター", ["IN_STOCK", "SOLD_OUT", "UNKNOWN"], default=["IN_STOCK", "SOLD_OUT", "UNKNOWN"], key="f_e")
            disp_e = df_e[df_e["_status"].isin(f_status_e)] if f_status_e else df_e
            
            if not disp_e.empty:
                edited_e_df = st.data_editor(
                    disp_e, 
                    column_config={
                        "_status": None, 
                        "ID": st.column_config.TextColumn("ID", disabled=True), 
                        "現在庫": st.column_config.NumberColumn("現在庫", disabled=True),
                        "在庫あり": st.column_config.CheckboxColumn("在庫あり", disabled=False),
                        "販売サイト": st.column_config.TextColumn("販売サイト", disabled=False),
                        "URL": st.column_config.LinkColumn("商品ページ")
                    }, 
                    width='stretch', 
                    hide_index=True, 
                    key="ed_e"
                )
                
                # --- 反映実行ボタン (テーブルの下に配置) ---
                st.markdown("---")
                btn_label = "🚀 eBayへ反映実行（チェック内容を同期）"
                if not auto_relist:
                    st.info("💡 在庫復活時の自動再出品がオフです。必要に応じてチェックを入れて反映してください。")
                
                if st.button(btn_label, type="primary", key="ebay_apply_btn"):
                    with st.spinner("eBayへ反映中..."):
                        try:
                            # 1. GUI の編集内容を CSV に保存
                            checker_apply = EbayStockChecker(dry_run=dry_run, auto_relist=auto_relist)
                            # 全データ行を渡してCSVを同期
                            checker_apply.save_gui_changes(edited_e_df.to_dict('records'))
                            
                            # 2. 保存された CSV 内容を eBay に反映
                            count = checker_apply.apply_csv_changes_to_ebay()
                            
                            st.success(f"🎉 {count} 件を eBay へ反映しました！")
                        except Exception as e:
                            st.error(f"エラー: {e}")
            else:
                st.info("💡 フィルター条件に一致する結果はありません。")
        else:
            st.info("在庫チェック後に結果が表示されます。")

import csv # eBay用
