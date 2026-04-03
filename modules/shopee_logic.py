import asyncio
import random
import re
import os
import glob
import shutil
import logging
import datetime
from urllib.parse import urlparse
from openpyxl import load_workbook
import openpyxl
from playwright.async_api import async_playwright, BrowserContext, Page

# ==========================================
# ログ設定 (統合プロジェクトのlogs/を使用)
# ==========================================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
log_dir = os.path.join(BASE_DIR, "logs")
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

process_log_path = os.path.join(log_dir, "process.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(process_log_path, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ==========================================
# URL復元関数
# ==========================================
def restore_url(sku_text: str) -> str:
    """SKUから有効なURLを復元する"""
    if not sku_text: return ""
    sku_str = str(sku_text).strip()
    if sku_str.startswith("http://") or sku_str.startswith("https://"): return sku_str

    target_domains = ["mercari.com", "amazon.co.jp", "yahoo.co.jp", "fril.jp", "yodobashi.com", "rakuten.co.jp", "shaddy.jp", "kikyoushingenmochi.com", "kurodama.co.jp", "biccamera.com"]
    if sku_str.startswith("amazon.co.jp"): sku_str = "www." + sku_str
    for domain in target_domains:
        if domain in sku_str: return f"https://{sku_str}"
    return ""

import openpyxl.worksheet.views
from openpyxl.descriptors.base import String

# Shopee固有のExcel読み込みエラー回避パッチ
openpyxl.worksheet.views.Pane.activePane = String(allow_none=True)
openpyxl.worksheet.views.Pane.state = String(allow_none=True)

# ==========================================
# 価格計算クラス
# ==========================================
class FinanceCalculator:
    @staticmethod
    def calculate_shopee_price(purchase_price: int, is_fleamarket: bool) -> int:
        if is_fleamarket: return purchase_price
        return round((purchase_price + 1800) / (1 - 0.4327))

# ==========================================
# スクレイピングの基底クラス
# ==========================================
class BaseScraper:
    is_fleamarket = False

    def __init__(self, context: BrowserContext):
        self.context = context

    @staticmethod
    def clean_price(text: str) -> int:
        if not text: return 0
        cleaned = re.sub(r'[^\d]', '', str(text))
        return int(cleaned) if cleaned else 0

    async def _fetch_and_parse(self, url: str) -> tuple[Page, any, str, int]:
        await asyncio.sleep(random.uniform(1.0, 3.0))
        page = await self.context.new_page()
        response = None
        status_code = 0
        title = ""
        error_msg = ""
        try:
            response = await page.goto(url, wait_until="domcontentloaded", timeout=45000)
            await page.wait_for_timeout(2000)
        except Exception as e:
            error_msg = str(e)
            
        try:
            title = await page.title()
            status_code = response.status if response else 0
        except Exception:
            pass

        content_for_log = ""
        try:
            content_for_log = await page.content()
        except:
            content_for_log = "Unable to get content"

        # ログパスを修正
        summary_log_path = os.path.join(log_dir, "scraping_history.txt")
        with open(summary_log_path, "a", encoding="utf-8") as f:
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{now}] URL: {url}\n")
            f.write(f" - Status Code: {status_code}\n")
            if error_msg: f.write(f" - Exception: {error_msg}\n")
            f.write(f" - Title: {title}\n")
            snip = content_for_log[:300].replace('\n', ' ')
            f.write(f" - HTML Snippet: {snip}...\n")
            f.write("--------------------------------------------------\n")

        if status_code == 0 or status_code >= 400:
            safe_url_name = re.sub(r'[^a-zA-Z0-9]', '_', url.replace('https://', ''))[:50]
            dump_html_path = os.path.join(log_dir, f"error_status{status_code}_{safe_url_name}.html")
            with open(dump_html_path, "w", encoding="utf-8") as f:
                f.write(f"<!-- URL: {url} -->\n<!-- Status Code: {status_code} -->\n<!-- Error: {error_msg} -->\n")
                f.write(content_for_log)

        return page, response, title, status_code

    async def check_stock(self, url: str) -> dict:
        raise NotImplementedError

# ==========================================
# サイト別スクレイピングクラス (eBayでも使えるようにクラス化)
# ==========================================
class YodobashiScraper(BaseScraper):
    async def check_stock(self, url: str) -> dict:
        from curl_cffi import requests
        title, status_code, error_msg, content, price = "", 0, "", "", 0
        try:
            await asyncio.sleep(random.uniform(1.0, 3.0))
            def _fetch(): return requests.get(url, impersonate="chrome110", timeout=30)
            response = await asyncio.to_thread(_fetch)
            status_code, content = response.status_code, response.text
            title_match = re.search(r'<title>(.*?)</title>', content, re.IGNORECASE | re.DOTALL)
            if title_match: title = title_match.group(1).strip()
            price_match = re.search(r'(?:id="js_scl_unitPrice"|class="[^"]*js-sales-price[^"]*")[^>]*>￥?([0-9,]+)', content)
            if price_match: price = self.clean_price(price_match.group(1))
        except Exception as e:
            status_code, error_msg = 0, str(e)
            
        if status_code == 0 or status_code >= 400:
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}

        stock_status = "SOLD_OUT"
        if "予定数の販売を終了しました" in content or "販売休止中です" in content: stock_status = "SOLD_OUT"
        elif "カートに入れる" in content or "予約する" in content or "お取り寄せ" in content: stock_status = "IN_STOCK"
        return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}

class AmazonScraper(BaseScraper):
    async def check_stock(self, url: str) -> dict:
        price, stock_status = 0, "UNKNOWN"
        try:
            page, response, title, status_code = await self._fetch_and_parse(url)
            if status_code == 0 or status_code >= 400:
                await page.close()
                return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
            
            content = await page.content()
            
            # コンテンツが小さすぎる場合は Bot 検知ページの可能性
            if len(content) < 10000:
                await page.close()
                return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
            
            # 価格取得
            try:
                price_loc = page.locator('.a-price-whole').first
                if await price_loc.count() > 0:
                    price_text = await price_loc.text_content(timeout=3000)
                    price = self.clean_price(price_text)
            except: pass
            
            # デフォルトは UNKNOWN（安全側）
            stock_status = "UNKNOWN"
            
            in_stock_keywords = ["カートに入れる", "今すぐ買う", "予約注文する", "add-to-cart-button", "買い物かごに入れる"]
            sold_out_keywords = ["現在在庫切れです", "現在お取り扱いできません", "在庫切れ", "この商品は現在お取り扱いしておりません", "Currently unavailable"]
            
            if any(kw in content for kw in sold_out_keywords):
                stock_status = "SOLD_OUT"
            if any(kw in content for kw in in_stock_keywords):
                stock_status = "IN_STOCK"
            
            await page.close()
        except Exception as e:
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
        return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}

class RakutenScraper(BaseScraper):
    async def check_stock(self, url: str) -> dict:
        from curl_cffi import requests
        from bs4 import BeautifulSoup
        title, status_code, error_msg, content, price, stock_status = "", 0, "", "", 0, "UNKNOWN"
        try:
            await asyncio.sleep(random.uniform(1.0, 3.0))
            def _fetch(): return requests.get(url, impersonate="chrome110", timeout=30)
            response = await asyncio.to_thread(_fetch)
            status_code, content = response.status_code, response.text
            soup = BeautifulSoup(content, 'html.parser')
            title_tag = soup.find('title')
            if title_tag: title = title_tag.text.strip()
            price_elem = soup.select_one('.price2')
            if price_elem: price = self.clean_price(price_elem.text)
            # デフォルトは UNKNOWN（判定できない場合は安全側）
            stock_status = "UNKNOWN"
            import re
            if "買い物かごに入れる" in content or "ご購入手続きへ" in content or "商品をかごに追加" in content or "カートに入れる" in content:
                stock_status = "IN_STOCK"
            elif re.search(r'"isEnableAddToCart"\s*:\s*true', content) or re.search(r'"purchaseCondition"\s*:\s*"enabled"', content) or re.search(r'"addtocart"\s*:\s*true', content):
                stock_status = "IN_STOCK"
            elif "在庫切れ" in content or "品切れ" in content or "販売中止" in content or "sold out" in content.lower():
                stock_status = "SOLD_OUT"
        except Exception as e:
            status_code, error_msg = 0, str(e)
            
        if status_code == 0 or status_code >= 400:
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
        return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}

class YahooShoppingScraper(BaseScraper):
    async def check_stock(self, url: str) -> dict:
        try:
            page, response, title, status_code = await self._fetch_and_parse(url)
            if status_code == 0 or status_code >= 400:
                await page.close()
                return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
            price = 0
            price_loc = page.locator('[data-e2e="itemPrice"], .ItemPrice_price__F3gE2')
            if await price_loc.count() > 0:
                try: price = self.clean_price(await price_loc.first.text_content(timeout=3000))
                except: pass
            content = await page.content()
            # デフォルトは UNKNOWN（判定できない場合は安全側）
            stock_status = "UNKNOWN"
            if "商品をカートに入れる" in content or "カートに入れる" in content or "予約する" in content:
                stock_status = "IN_STOCK"
            elif "在庫切れ" in content or "品切れ" in content or "販売中止" in content or "買えない" in content:
                stock_status = "SOLD_OUT"
            await page.close()
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}
        except Exception as e:
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN", "error": str(e)}

class MercariScraper(BaseScraper):
    is_fleamarket = True
    async def check_stock(self, url: str) -> dict:
        try:
            page, response, title, status_code = await self._fetch_and_parse(url)
            if status_code == 0 or status_code >= 400:
                await page.close()
                return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
            try: await page.locator('[data-testid="price"], [data-testid="checkout-button"]').first.wait_for(state="attached", timeout=10000)
            except: pass
            await page.wait_for_timeout(1000)
            price = 0
            price_loc = page.locator('[data-testid="price"]')
            if await price_loc.count() > 0:
                try: price = self.clean_price(await price_loc.first.text_content(timeout=3000))
                except: pass
            stock_status = "SOLD_OUT"
            buy_btns = page.locator('text="購入手続きへ"')
            btn_count = await buy_btns.count()
            for i in range(btn_count):
                btn = buy_btns.nth(i)
                if await btn.is_visible():
                    is_disabled = await btn.evaluate('el => el.disabled || el.hasAttribute("disabled") || el.getAttribute("aria-disabled") === "true"')
                    if not is_disabled:
                        stock_status = "IN_STOCK"
                        break
            try:
                sold_badge = page.locator('mer-item-thumbnail[sticker="sold"]')
                if await sold_badge.count() > 0: stock_status = "SOLD_OUT"
            except: pass
            await page.close()
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}
        except Exception as e:
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN", "error": str(e)}

class YahooAuctionScraper(BaseScraper):
    is_fleamarket = True
    async def check_stock(self, url: str) -> dict:
        try:
            page, response, title, status_code = await self._fetch_and_parse(url)
            if status_code == 0 or status_code >= 400:
                await page.close()
                return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
            price = 0
            price_loc = page.locator('.Price__value')
            if await price_loc.count() > 0:
                try: price = self.clean_price(await price_loc.first.text_content(timeout=3000))
                except: pass
            content = await page.content()
            stock_status = "SOLD_OUT"
            if "入札する" in content or "今すぐ落札する" in content or "購入手続きへ" in content: stock_status = "IN_STOCK"
            await page.close()
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}
        except Exception as e:
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN", "error": str(e)}

class YahooFleamarketScraper(BaseScraper):
    is_fleamarket = True
    async def check_stock(self, url: str) -> dict:
        try:
            page, response, title, status_code = await self._fetch_and_parse(url)
            if status_code == 0 or status_code >= 400:
                await page.close()
                return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
            try: await page.locator('[class*="Price__value"], #item_buy_button').first.wait_for(state="attached", timeout=10000)
            except: pass
            await page.wait_for_timeout(1000)
            price = 0
            price_loc = page.locator('[class*="Price__value"]')
            if await price_loc.count() > 0:
                try: price = self.clean_price(await price_loc.first.text_content(timeout=3000))
                except: pass
            stock_status = "SOLD_OUT"
            buy_btns = page.locator('text="購入手続きへ"')
            btn_count = await buy_btns.count()
            for i in range(btn_count):
                btn = buy_btns.nth(i)
                if await btn.is_visible():
                    is_disabled = await btn.evaluate('el => el.disabled || el.hasAttribute("disabled") || el.getAttribute("aria-disabled") === "true"')
                    if not is_disabled:
                        stock_status = "IN_STOCK"
                        break
            await page.close()
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}
        except Exception as e:
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN", "error": str(e)}

class RakumaScraper(BaseScraper):
    is_fleamarket = True
    async def check_stock(self, url: str) -> dict:
        try:
            page, response, title, status_code = await self._fetch_and_parse(url)
            if status_code == 0 or status_code >= 400:
                await page.close()
                return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
            price = 0
            price_loc = page.locator('.item__price')
            if await price_loc.count() > 0:
                try: price = self.clean_price(await price_loc.first.text_content(timeout=3000))
                except: pass
            content = await page.content()
            stock_status = "SOLD_OUT"
            if "購入に進む" in content or "購入する" in content or "カートに入れる" in content: stock_status = "IN_STOCK"
            await page.close()
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}
        except Exception as e:
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN", "error": str(e)}

class KikyouScraper(BaseScraper):
    async def check_stock(self, url: str) -> dict:
        from curl_cffi import requests
        from bs4 import BeautifulSoup
        title, status_code, error_msg, content, price, stock_status = "", 0, "", "", 0, "UNKNOWN"
        try:
            await asyncio.sleep(random.uniform(1.0, 3.0))
            def _fetch(): return requests.get(url, impersonate="chrome110", timeout=30)
            response = await asyncio.to_thread(_fetch)
            status_code, content = response.status_code, response.text
            soup = BeautifulSoup(content, 'html.parser')
            title_tag = soup.find('title')
            if title_tag: title = title_tag.text.strip()
            stock_status = "SOLD_OUT"
            if "カートに入れる" in content or "カゴに入れる" in content or "sysCartInButton" in content: stock_status = "IN_STOCK"
        except Exception as e:
            status_code, error_msg = 0, str(e)
        if status_code == 0 or status_code >= 400: return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
        return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}

class ShaddyScraper(BaseScraper):
    async def check_stock(self, url: str) -> dict:
        from curl_cffi import requests
        from bs4 import BeautifulSoup
        title, status_code, error_msg, content, price, stock_status = "", 0, "", "", 0, "UNKNOWN"
        try:
            await asyncio.sleep(random.uniform(1.0, 3.0))
            def _fetch(): return requests.get(url, impersonate="chrome110", timeout=30)
            response = await asyncio.to_thread(_fetch)
            status_code, content = response.status_code, response.text
            soup = BeautifulSoup(content, 'html.parser')
            title_tag = soup.find('title')
            if title_tag: title = title_tag.text.strip()
            stock_status = "SOLD_OUT"
            if "カート" in content or "カゴ" in content: stock_status = "IN_STOCK"
            if "在庫切れ" in content: stock_status = "SOLD_OUT"
        except Exception as e:
            status_code, error_msg = 0, str(e)
        if status_code == 0 or status_code >= 400: return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
        return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}

class KurodamaScraper(BaseScraper):
    async def check_stock(self, url: str) -> dict:
        from curl_cffi import requests
        from bs4 import BeautifulSoup
        title, status_code, error_msg, content, price, stock_status = "", 0, "", "", 0, "UNKNOWN"
        try:
            await asyncio.sleep(random.uniform(1.0, 3.0))
            def _fetch(): return requests.get(url, impersonate="chrome110", timeout=30)
            response = await asyncio.to_thread(_fetch)
            status_code, content = response.status_code, response.text
            soup = BeautifulSoup(content, 'html.parser')
            title_tag = soup.find('title')
            if title_tag: title = title_tag.text.strip()
            stock_status = "SOLD_OUT"
            if "カートに入れる" in content or "カゴに入れる" in content: stock_status = "IN_STOCK"
            if "sold out" in content.lower() or "売切れ" in content or "品切れ" in content: stock_status = "SOLD_OUT"
        except Exception as e:
            status_code, error_msg = 0, str(e)
        if status_code == 0 or status_code >= 400: return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
        return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": price, "status": stock_status}

class GenericScraper(BaseScraper):
    async def check_stock(self, url: str) -> dict:
        try:
            page, response, title, status_code = await self._fetch_and_parse(url)
            await page.close()
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN"}
        except Exception as e:
            return {"url": url, "success": True, "is_fleamarket": self.is_fleamarket, "price": 0, "status": "UNKNOWN", "error": str(e)}

class ScraperFactory:
    @staticmethod
    def get_scraper(url: str, context: BrowserContext) -> BaseScraper:
        domain = url.lower()
        if "yodobashi.com" in domain: return YodobashiScraper(context)
        elif "amazon.co.jp" in domain: return AmazonScraper(context)
        elif "rakuten.co.jp" in domain: return RakutenScraper(context)
        elif "shopping.yahoo.co.jp" in domain: return YahooShoppingScraper(context)
        elif "auctions.yahoo.co.jp" in domain or "page.auctions.yahoo.co.jp" in domain: return YahooAuctionScraper(context)
        elif "mercari" in domain: return MercariScraper(context)
        elif "paypayfleamarket.yahoo.co.jp" in domain or "fleamarket.yahoo.co.jp" in domain: return YahooFleamarketScraper(context)
        elif "fril.jp" in domain: return RakumaScraper(context)
        elif "shaddy.jp" in domain: return ShaddyScraper(context)
        elif "kikyoushingenmochi.com" in domain: return KikyouScraper(context)
        elif "kurodama.co.jp" in domain: return KurodamaScraper(context)
        else: return GenericScraper(context)

# ==========================================
# Shopee 自動化機能 (DL / UL)
# ==========================================
class ShopeeLoginRequiredError(Exception):
    """Shopeeのログインが必要な場合にスローされる例外"""
    pass

async def auto_download_shopee(page: Page, download_dir: str) -> str:
    if not os.path.exists(download_dir): os.makedirs(download_dir)
    await page.bring_to_front()
    await page.goto("https://seller.shopee.sg/portal/product-mass/mass-update/download", wait_until="domcontentloaded", timeout=60000)
    
    # ログインページにリダイレクトされたかチェック
    if "login" in page.url or await page.locator("input[name='loginKey']").count() > 0:
        raise ShopeeLoginRequiredError("Shopeeへのログインが必要です。")

    try:
        await page.wait_for_selector("text='Sales Info'", timeout=30000)
    except Exception:
        if "login" in page.url:
            raise ShopeeLoginRequiredError("Shopeeへのログインが必要です。")
        raise
    
    await page.locator("label").filter(has_text="Sales Info").click()
    await page.click("button:has-text('Generate')")
    await asyncio.sleep(15)
    async with page.expect_download(timeout=60000) as dl_info:
        await page.locator("button:has-text('Download')").first.click()
    download = await dl_info.value
    file_path = os.path.join(download_dir, download.suggested_filename)
    await download.save_as(file_path)
    return file_path

async def auto_upload_shopee(page: Page, file_path: str):
    await page.goto("https://seller.shopee.sg/portal/product-mass/mass-update/upload", wait_until="domcontentloaded", timeout=60000)
    if "login" in page.url: return False
    file_input_selector = "input[type='file']"
    await page.wait_for_selector(file_input_selector, state="attached", timeout=30000)
    await page.locator(file_input_selector).first.set_input_files(file_path)
    await page.wait_for_selector("text='Processed'", timeout=60000)
    await page.wait_for_timeout(5000)
    return True

# ==========================================
# コアロジッククラス
# ==========================================
class ShopeeStockChecker:
    def __init__(self, headless=True):
        self.headless = headless
        self.user_data_dir = os.path.join(BASE_DIR, "shopee_browser_data")
        self.download_dir = os.path.join(BASE_DIR, "downloads")
        self.ready_to_upload_dir = os.path.join(BASE_DIR, "Ready_to_Upload")
        self.max_concurrent_tasks = 8

    async def run_full_cycle(self, progress_callback=None, skip_price_update=False):
        async with async_playwright() as p:
            browser = await p.chromium.launch_persistent_context(
                user_data_dir=self.user_data_dir,
                headless=self.headless,
                args=["--disable-blink-features=AutomationControlled"]
            )
            try:
                page = browser.pages[0] if browser.pages else await browser.new_page()
                if progress_callback: progress_callback("Shopeeから最新データを取得中...", 0, 100)
                latest_file = await auto_download_shopee(page, self.download_dir)
                wb = openpyxl.load_workbook(latest_file)
                ws = wb.active
                tasks = self._extract_tasks(ws)
                total = len(tasks)
                if progress_callback: progress_callback(f"チェック対象: {total} 件発見", 0, total)
                domain_semaphores = {}
                global_semaphore = asyncio.Semaphore(self.max_concurrent_tasks)
                url_cache = {}
                processed_results = []
                def get_domain_semaphore(url_str):
                    domain = urlparse(url_str).netloc
                    if domain not in domain_semaphores: domain_semaphores[domain] = asyncio.Semaphore(1)
                    return domain_semaphores[domain]
                async def fetch_item(task):
                    url = task["url"]
                    if url not in url_cache:
                        async def _do():
                            async with global_semaphore, get_domain_semaphore(url):
                                return await ScraperFactory.get_scraper(url, browser).check_stock(url)
                        url_cache[url] = asyncio.create_task(_do())
                    res = await url_cache[url]
                    item_res = {"row_num": task["row_num"], "display_id": task["display_id"], "result": res}
                    processed_results.append(item_res)
                    if progress_callback: progress_callback(f"進捗: {len(processed_results)}/{total}", len(processed_results), total, processed_results)
                    return item_res
                await asyncio.gather(*(fetch_item(t) for t in tasks))
                if progress_callback: progress_callback("結果をExcelに保存中...", total, total, processed_results)
                output_path, summary = self._save_results(wb, ws, processed_results, skip_price_update=skip_price_update)
                return output_path, summary, page
            finally:
                if not progress_callback: await browser.close()

    def _extract_tasks(self, ws):
        sku_cols = []
        product_id_col = None
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    v = cell.value.strip().lower()
                    if v in ("sku", "parent sku", "variant sku"): sku_cols.append(cell.column)
                    elif v == "product id": product_id_col = cell.column
        tasks = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            url, cell_sku, sku_cell = "", "", None
            for idx in sku_cols:
                c = row[idx-1]
                if c and c.value:
                    v = str(c.value).strip()
                    if v.lower() in ("sku", "parent sku", "variant sku"): continue
                    restored = restore_url(v)
                    if restored.startswith("http"):
                        url, cell_sku, sku_cell = restored, v, c
                        break
            if not url: continue
            pid = str(row[product_id_col-1].value) if product_id_col and row[product_id_col-1].value else ""
            tasks.append({"row_num": sku_cell.row, "url": url, "display_id": pid if pid else cell_sku})
        return tasks

    def _save_results(self, wb, ws, results, skip_price_update=False):
        stock_col, price_col = None, None
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value == "Stock": stock_col = cell.column
                elif cell.value == "Price": price_col = cell.column
        stats = {"total": 0, "in_stock": 0, "sold_out": 0, "unknown": 0, "updated": 0}
        for item in results:
            res = item["result"]
            row_num = item["row_num"]
            if res.get("success"):
                status = res["status"]
                price = res["price"]
                stats["total"] += 1
                if status == "SOLD_OUT":
                    stats["sold_out"] += 1
                    if stock_col: ws.cell(row=row_num, column=stock_col).value = 0
                    stats["updated"] += 1
                elif status == "IN_STOCK":
                    stats["in_stock"] += 1
                    if stock_col: ws.cell(row=row_num, column=stock_col).value = 1
                    if not skip_price_update and not res["is_fleamarket"] and price_col:
                        ws.cell(row=row_num, column=price_col).value = FinanceCalculator.calculate_shopee_price(price, False)
                    stats["updated"] += 1
                else: stats["unknown"] += 1
        if not os.path.exists(self.ready_to_upload_dir): os.makedirs(self.ready_to_upload_dir)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = os.path.join(self.ready_to_upload_dir, f"mass_update_ready_{ts}.xlsx")
        wb.save(output_filename)
        return output_filename, stats

    def save_manual_results(self, wb, ws, edited_df_as_dict, skip_price_update=False):
        stock_col, price_col = None, None
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value == "Stock": stock_col = cell.column
                elif cell.value == "Price": price_col = cell.column

        stats = {"total": len(edited_df_as_dict), "updated": 0}
        for item in edited_df_as_dict:
            row_num = item.get("row_num")
            if not row_num: continue
            
            is_in_stock = item.get("在庫あり")
            if is_in_stock is None:
                is_in_stock = (item.get("ステータス") == "IN_STOCK")

            manual_shopee_price = item.get("出品価格(直接)")
            purchase_price = item.get("仕入価格")
            is_fleamarket = item.get("is_fleamarket", False)

            if not is_in_stock:
                if stock_col: ws.cell(row=row_num, column=stock_col).value = 0
            else:
                if stock_col: ws.cell(row=row_num, column=stock_col).value = 1
                
                if not skip_price_update:
                    final_price = 0
                    if manual_shopee_price and manual_shopee_price > 0:
                        final_price = manual_shopee_price
                    elif purchase_price and purchase_price > 0:
                        final_price = FinanceCalculator.calculate_shopee_price(purchase_price, is_fleamarket)
                    if price_col and final_price > 0:
                        ws.cell(row=row_num, column=price_col).value = final_price
            stats["updated"] += 1

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = os.path.join(self.ready_to_upload_dir, f"mass_update_manual_{ts}.xlsx")
        wb.save(output_filename)
        return output_filename, stats
